import csv
import io
import os
from abc import ABC, abstractmethod
from enum import Enum

from openpyxl import load_workbook

from extensions.deployment.exceptions import HumaOptionLibraryInvalidDataException
from sdk.storage.model.file_storage import FileStorage
from sdk.storage.use_case.storage_request_objects import DownloadFileRequestObjectV1
from sdk.storage.use_case.storage_use_cases import DownloadFileUseCaseV1


class MetadataFields:
    DEPLOYMENT_ID = "deploymentId"
    LIBRARY_ID = "libraryId"
    DESCRIPTION = "description"
    LIST_NAME = "listName"
    LIST_COUNT = "listCount"


class LibraryType(Enum):
    HUMA_IMAGE_LIBRARY = "huma_image_library"
    HUMA_OPTIONS_LIBRARY = "huma_options_library"
    DEPLOYMENT_OPTIONS_LIBRARY = "deployment_options_library"


STANDARD_LIBRARIES = [
    LibraryType.HUMA_IMAGE_LIBRARY.value,
    LibraryType.HUMA_OPTIONS_LIBRARY.value,
]


class MetadataCreator(ABC):
    @staticmethod
    @abstractmethod
    def create_metadata(file: FileStorage) -> dict:
        raise NotImplementedError

    @staticmethod
    def _file_content(file: FileStorage):
        req = DownloadFileRequestObjectV1(fileId=file.id)
        rsp = DownloadFileUseCaseV1().execute(req)
        return rsp.value.content


class HumaImageLibraryMetadataCreator(MetadataCreator):
    @staticmethod
    def create_metadata(file: FileStorage) -> dict:
        description = os.path.splitext(file.fileName)[0] if file.fileName else None
        return {MetadataFields.DESCRIPTION: description}


class OptionsLibraryMetadataCreator(MetadataCreator):
    @staticmethod
    def create_metadata(file: FileStorage) -> dict:
        list_name = os.path.splitext(file.fileName)[0] if file.fileName else None
        try:
            return {
                MetadataFields.LIST_NAME: list_name,
                MetadataFields.LIST_COUNT: OptionsLibraryMetadataCreator.file_list_size(
                    file
                ),
            }
        except Exception:
            raise HumaOptionLibraryInvalidDataException()

    @staticmethod
    def file_list_size(file: FileStorage):
        title_column = 1
        content = MetadataCreator._file_content(file)
        if file.contentType.lower() == "text/csv":
            all_rows = len(list(csv.reader(io.TextIOWrapper(content, "utf-8"))))
            return all_rows - title_column

        excel_types = [
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        ]
        if file.contentType.lower() in excel_types:
            workbook = load_workbook(filename=content)
            return workbook[workbook.sheetnames[0]].max_row - title_column


class HumaOptionsLibraryMetadataCreator(OptionsLibraryMetadataCreator):
    pass


class DeploymentOptionsLibraryMetadataCreator(OptionsLibraryMetadataCreator):
    pass


class FileLibraryMetadataFactory:
    library_metadata_creators = {
        LibraryType.HUMA_IMAGE_LIBRARY.value: HumaImageLibraryMetadataCreator,
        LibraryType.HUMA_OPTIONS_LIBRARY.value: HumaOptionsLibraryMetadataCreator,
        LibraryType.DEPLOYMENT_OPTIONS_LIBRARY.value: DeploymentOptionsLibraryMetadataCreator,
    }

    @staticmethod
    def get_metadata_creator(libraryId: str) -> MetadataCreator:
        return FileLibraryMetadataFactory.library_metadata_creators.get(libraryId)
